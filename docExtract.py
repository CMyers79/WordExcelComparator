#  This program extracts cost and energy savings tables from Word .docx files and searches for user-provided
#  labels that correspond to the cost and energy savings in eProjectBuilder Schedule 4 columns Y and AE.
#  The $/yr and MBTU/yr numbers are then copied to the Comparator.xlsx template and compared to the Schedule 4 values.
#  For building-specific ECM tables, the program searches for the building name in the first row of the table.


import openpyxl
from docx import Document


# helper function returns boolean for string contains int or float
def is_float(string):
    try:
        float(string.replace(",", ""))
        return True
    except ValueError:
        return False


# The docExtractor class should be accessed by the workflow in comparator.py.
class DocExtractor:
    # construct an instance of docExtractor with doc_type 'docx', 'xlsx', or 'epb'
    def __init__(self, doc_type):
        self.data = None
        self.doc_name = None
        self.ecms = [[], []]
        self.output = openpyxl.load_workbook('Comparator.xlsx')
        self.doc_type = doc_type

    # load file at filepath passed in
    def load(self, filepath):
        # load file into self.data
        if self.doc_type == 'epb' or self.doc_type == 'xlsx':
            self.data = openpyxl.load_workbook(filepath, data_only=True)
        elif self.doc_type == 'docx':
            self.data = Document(filepath)
            if filepath == 'Vol1.docx':
                self.doc_name = 'Vol1'
            elif filepath == 'Vol2.docx':
                self.doc_name = 'Vol2'

    # extract table data from self.data by self.doc_type
    def extract(self, ecm_list):
        # load output template for population
        comparator = self.output.get_sheet_by_name('comparator')
        ecm = []
        cost = []
        mbtu = []
        include = []
        baseline = [[], [], [], []]
        savings = [[], [], [], []]
        cost_savings = [[], [], [], []]
        m_v_option = []
        o_m_cost_savings = []
        other_cost_savings = []
        price = []
        spb = []

        # TODO refactor to use class inheritance
        # pull data from columns Y and AE in the epb calculating template
        if self.doc_type == 'epb':
            schedule_4 = self.data.get_sheet_by_name('Sch4-Cost Savings by ECM')
            summary = self.data.get_sheet_by_name('Summary Schedule')

            # populate identifier fields in comparator
            comparator.cell(row=2, column=2, value=summary.cell(row=16, column=3).value)
            # no reliable data source for ESPC Phase
            comparator.cell(row=4, column=2, value=summary.cell(row=9, column=3).value)

            # populate arrays with 250 ECM rows plus the totals
            for i in range(251):
                ecm.append(schedule_4.cell(row=i+8, column=1).value)
                mbtu.append(schedule_4.cell(row=i+8, column=25).value)
                cost.append(schedule_4.cell(row=i+8, column=31).value)
                m_v_option.append(schedule_4.cell(row=i + 8, column=4).value)
                o_m_cost_savings.append(schedule_4.cell(row=i + 8, column=29).value)
                other_cost_savings.append(schedule_4.cell(row=i + 8, column=30).value)
                price.append(schedule_4.cell(row=i + 8, column=32).value)
                spb.append(schedule_4.cell(row=i + 8, column=33).value)
                baseline[0].append(schedule_4.cell(row=i+8, column=5).value)  # kWh
                baseline[1].append(schedule_4.cell(row=i+8, column=6).value)  # kW
                baseline[2].append(schedule_4.cell(row=i+8, column=7).value)  # MBtu
                baseline[3].append(schedule_4.cell(row=i+8, column=10).value)  # kGal
                savings[0].append(schedule_4.cell(row=i + 8, column=15).value)  # kWh
                savings[1].append(schedule_4.cell(row=i + 8, column=17).value)  # kW
                savings[2].append(schedule_4.cell(row=i + 8, column=19).value)  # MBtu
                savings[3].append(schedule_4.cell(row=i + 8, column=27).value)  # kGal
                cost_savings[0].append(schedule_4.cell(row=i + 8, column=16).value)  # kWh
                cost_savings[1].append(schedule_4.cell(row=i + 8, column=18).value)  # kW
                cost_savings[2].append(schedule_4.cell(row=i + 8, column=20).value)  # MBtu
                cost_savings[3].append(schedule_4.cell(row=i + 8, column=28).value)  # kGal

                # add non-blank row indices to include
                if ecm[i] not in (None, ''):
                    include.append(i)

            # remove 'TOTALS' from ecm
            ecm.pop()

            # copy non-blank ECM rows to Comparator.xlsx
            for i in range(len(include) - 1):
                comparator.cell(row=18 + i, column=1, value=ecm[include[i]])
                comparator.cell(row=18 + i, column=2, value=cost[include[i]])
                comparator.cell(row=18 + i, column=9, value=mbtu[include[i]])
                comparator.cell(row=18 + i, column=16, value=price[include[i]])
                comparator.cell(row=18 + i, column=23, value=baseline[0][include[i]])
                comparator.cell(row=18 + i, column=30, value=baseline[1][include[i]])
                comparator.cell(row=18 + i, column=37, value=baseline[2][include[i]])
                comparator.cell(row=18 + i, column=44, value=baseline[3][include[i]])
                comparator.cell(row=18 + i, column=51, value=savings[0][include[i]])
                comparator.cell(row=18 + i, column=58, value=savings[1][include[i]])
                comparator.cell(row=18 + i, column=65, value=savings[2][include[i]])
                comparator.cell(row=18 + i, column=72, value=savings[3][include[i]])
                comparator.cell(row=18 + i, column=79, value=cost_savings[0][include[i]])
                comparator.cell(row=18 + i, column=86, value=cost_savings[1][include[i]])
                comparator.cell(row=18 + i, column=93, value=cost_savings[2][include[i]])
                comparator.cell(row=18 + i, column=100, value=cost_savings[3][include[i]])
                comparator.cell(row=18 + i, column=107, value=m_v_option[include[i]])
                comparator.cell(row=18 + i, column=114, value=o_m_cost_savings[include[i]])
                comparator.cell(row=18 + i, column=121, value=other_cost_savings[include[i]])
                comparator.cell(row=18 + i, column=128, value=spb[include[i]])

            # copy totals to Comparator.xlsx
            comparator.cell(row=16, column=2, value=cost[250])
            comparator.cell(row=16, column=9, value=mbtu[250])
            comparator.cell(row=18 + len(include) - 1, column=2, value=cost[250])
            comparator.cell(row=18 + len(include) - 1, column=9, value=mbtu[250])

            # populate ecm_split list with building names and ecm numbers
            for title in [ecm_title for ecm_title in ecm if ecm_title not in ["", None]]:
                # the ecm number may end in a letter, so look at the penultimate character
                i = len(title) - 2
                # find the last non-decimal character, this is the end of the building name
                while title[i] in "0123456789.":
                    i -= 1
                # add the building name to ecms[0]
                self.ecms[0].append(title[:i + 1].strip().replace('.', ''))

                # find the first digit after the building name, the start of the ecm number
                i = 0
                while title[i] not in "012456789":
                    i += 1
                # add all building ecm numbers to ecms[1]
                self.ecms[1].append(title[i:].strip())

        # pull data from the savings per ECM template
        elif self.doc_type == 'xlsx':
            ecm_savings = self.data.get_sheet_by_name('Sheet1')

            # pull data by ECM number starting with B3, copy directly to comparator
            row = 3
            while ecm_savings.cell(row=row, column=2).value not in (None, ''):
                comparator.cell(row=15 + row, column=135, value=ecm_savings.cell(row=row, column=2).value)
                comparator.cell(row=15 + row, column=136, value=ecm_savings.cell(row=row, column=9).value)
                row += 1

            # copy total to Comparator.xlsx
            comparator.cell(row=16, column=136, value=ecm_savings.cell(row=row, column=9).value)

        # search for table data within Word docx file
        elif self.doc_type == 'docx':
            # populate list with alias inputs
            aliases = [[comparator.cell(row=11, column=3 + i).value for i in range(5)],
                       [comparator.cell(row=11, column=10 + j).value for j in range(5)]]
            for data_offset in range(21):
                aliases.append([comparator.cell(row=13, column=3 + 7 * data_offset + alias_offset).value for alias_offset in range(5)])
            print("aliases ", aliases)
            print("Extracting:")
            # iterate through each table in the Word docx (xml) and load the cell values into an array
            for q, table in enumerate(self.data.tables):
                cell_text = []
                for row in table.rows:
                    cell_text.append([cell.text for cell in row.cells])

                # search table for aliases from Comparator.xlsx
                for i, row in enumerate(cell_text):
                    for j, cell in enumerate(row):
                        # search each table cell for each alias
                        for k, alias_5 in enumerate(aliases):
                            # if total $ savings is found:
                            for c in range(5):
                                if cell == alias_5[c] and cell not in ["", None] and k == 0:
                                    print("Table " + str(q) + " " + alias_5[c])
                                    # pull the numerical value in the cell to either the right or bottom of the alias
                                    if i < len(cell_text) - 1 and is_float(cell_text[i + 1][j].replace("$", "")):
                                        comparator.cell(row=16,
                                                        column=3 + c,
                                                        value=float(cell_text[i+1][j].replace(",", "").replace("$", "")))

                                    elif j < len(cell_text[i]) - 1 and is_float(cell_text[i][j + 1].replace("$", "")):
                                        comparator.cell(row=16,
                                                        column=3 + c,
                                                        value=float(cell_text[i][j + 1].replace(",", "").replace("$", "")))

                                # if total MBTU is found
                                elif cell == alias_5[c] and cell not in ["", None] and k == 1:
                                    print("Table " + str(q) + " " + alias_5[c])
                                    # pull the numerical value in the cell to either the right or bottom of the alias
                                    if i < len(cell_text) - 1 and is_float(cell_text[i + 1][j]):
                                        comparator.cell(row=16,
                                                        column=10 + c,
                                                        value=float(cell_text[i+1][j].replace(",", "")))

                                    elif j < len(cell_text[i]) - 1 and is_float(cell_text[i][j + 1]):
                                        comparator.cell(row=16,
                                                        column=10 + c,
                                                        value=float(cell_text[i][j + 1].replace(",", "")))

                                # if ECM alias match is found
                                elif cell == alias_5[c] and cell not in ["", None]:
                                    print("Table " + str(q) + " " + alias_5[c])
                                    # find the building name in the first row
                                    n = 0
                                    while n < len(cell_text[0]) - 1 and cell_text[0][n] in ["", None]:
                                        n += 1
                                    b_name = cell_text[0][n]
                                    b_offset = 0
                                    found_offset = False
                                    b_ecms = 0

                                    # determine the offset of the building in ecm_list and how many ecms at the building
                                    for m, name in enumerate(ecm_list[0]):
                                        # must ignore the last word in the building name because it may be 'TC' or 'ECM'
                                        # unless building name is one word
                                        if len(name.split(" ")) > 1:
                                            split_name = " ".join(name.split(" ")[:-1])
                                        else:
                                            split_name = name

                                        if split_name in b_name or b_name in split_name:
                                            # set b_offset to the first row with matching building name
                                            if not found_offset:
                                                b_offset = m
                                                found_offset = True

                                            b_ecms += 1

                                    # pull the numerical values in the cells to either the right or bottom of the alias
                                    m = i
                                    n = j
                                    # skip blank cells and check the first populated cell
                                    while m < len(cell_text) - 1 and cell_text[m + 1][j] == '\xa0':
                                        m += 1
                                    while n < len(cell_text[i]) - 1 and cell_text[i][n + 1] == '\xa0':
                                        n += 1

                                    if is_float(cell_text[m + 1][j].replace(",", "").replace("$", "")):
                                        p = i

                                        while p < len(cell_text) - 1 and (
                                                is_float(cell_text[p + 1][j].replace(",", "").replace("$", "")) or cell_text[p + 1][j] == '\xa0') \
                                                and p - i < b_ecms:
                                            if cell_text[p + 1][j] == '\xa0':
                                                p += 1
                                                continue
                                            else:
                                                comparator.cell(row=18 + b_offset + p - i,
                                                                column=3 + 7 * k + c - 14,
                                                                value=float(cell_text[p + 1][j].replace(",", "").replace("$", "")))
                                            p += 1

                                    elif is_float(cell_text[i][n + 1].replace(",", "").replace("$", "")):
                                        p = j

                                        while p < len(cell_text[i]) - 1 and (
                                                is_float(cell_text[i][p + 1].replace(",", "").replace("$", "")) or cell_text[i][p + 1] == '\xa0') \
                                                and p - j < b_ecms:
                                            if cell_text[i][p + 1] == '\xa0':
                                                p += 1
                                                continue
                                            else:
                                                comparator.cell(row=18 + b_offset + p - j,
                                                                column=3 + 7 * k + c - 14,
                                                                value=float(cell_text[i][p + 1].replace(",", "").replace("$", "")))

        self.output.save('Comparator.xlsx')
