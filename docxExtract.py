#  This program extracts cost and energy savings tables from Word .docx files and searches for user-provided
#  labels that correspond to the cost and energy savings in eProjectBuilder Schedule 4 columns Y and AE.
#  The $/yr and MBTU/yr numbers are then copied to the Comparator.xlsx template and compared to the Schedule 4 values.


import openpyxl
from docx import Document


#  The docExtractor class should be accessed by the workflow in comparator.py.
class docExtractor:

    #  construct an instance of docExtractor with doc_type 'docx', 'xlsx', or 'epb'
    def __init__(self, doc_type):
        self.data = None
        self.output = openpyxl.load_workbook('Comparator.xlsx')
        self.doc_type = doc_type

    # extract table data from the filepath
    def load(self, filepath):
        #  load file into self.data
        if self.doc_type == 'epb' or self.doc_type == 'xlsx':
            self.data = openpyxl.load_workbook(filepath, data_only=True)
        elif self.doc_type == 'docx':
            self.data = Document(filepath)

    def extract(self):
        # pull data from columns Y and AE in the epb calculating template
        if self.doc_type == 'epb':
            schedule_4 = self.data.get_sheet_by_name('Sch4-Cost Savings by ECM')
            comparator = self.output.get_sheet_by_name('comparator')
            ecm = []
            mbtu = []
            cost = []
            include = []
            # populate arrays with 250 ECM rows plus the total
            for i in range(251):
                ecm.append(schedule_4.cell(row=i+8, column=0).value)
                mbtu.append(schedule_4.cell(row=i+8, column=24).value)
                cost.append(schedule_4.cell(row=i+8, column=30).value)
                # add non-blank row indices to include
                if ecm[i] not in (None, ''):
                    include.append(i)
            # copy ECM rows to Comparator.xlsx
            for i in range(len(include) - 1):
                comparator.cell(row=9 + i, column=0, value=ecm[include[i]])
                comparator.cell(row=9 + i, column=1, value=cost[include[i]])
                comparator.cell(row=9 + i, column=6, value=mbtu[include[i]])

            # copy totals to Comparator.xlsx
            comparator.cell(row=7, column=1, value=cost[250])
            comparator.cell(row=7, column=6, value=mbtu[250])

        # pull data from the savings per ECM template




